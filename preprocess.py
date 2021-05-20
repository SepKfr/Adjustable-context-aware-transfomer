import numpy as np
import pandas as pd
import openpyxl
import torch
import pickle
from sklearn.preprocessing import StandardScaler
import argparse
import pywt
import random
random.seed(21)
torch.manual_seed(21)
np.random.seed(21)


class Scaler:
    def __init__(self, site):
        self.site = site
        self.scalers = dict()

    def add_scaler(self, f_n, set_dat, scaler):
        self.scalers[(set_dat, f_n)] = scaler


class Data:
    def __init__(self, site_data, params, n_features):

        self.scalers = list()
        self.sites_data = site_data
        self.ts = params.max_length
        self.add_wave = True if params.add_wave == "True" else False
        self.n_seasons = 4
        self.moving_averages = [4, 8, 16, 32]
        self.n_moving_average = 0
        self.wavelets = ['db1', 'db2', 'db3']
        self.n_wavelets = 0
        self.nf = 1
        self.in_seq_len = params.in_seq_len
        self.out_seq_len = params.out_seq_len

        self.train_ts = int(self.ts * params.train_percent)
        self.valid_ts = int((self.ts - self.train_ts) * 0.5)
        self.test_ts = self.valid_ts

        self.train_ln = self.get_length(self.train_ts)
        self.valid_ln = self.get_length(self.valid_ts)
        self.test_ln = self.get_length(self.test_ts)

        self.train_x = torch.zeros(self.train_ln, self.in_seq_len, self.nf)
        self.train_y = torch.zeros(self.train_ln, self.out_seq_len, 1)
        self.valid_x = torch.zeros(self.valid_ln, self.in_seq_len, self.nf)
        self.valid_y = torch.zeros(self.valid_ln, self.out_seq_len, 1)
        '''self.test_x = torch.zeros(self.test_ln, self.in_seq_len, self.nf)
        self.test_y = torch.zeros(self.test_ln, self.out_seq_len, 1)'''

        self.scalers = list()

        for abr, df_site in self.sites_data.items():

            scaler_per_site = Scaler(abr)
            self.scalers.append(scaler_per_site)
            df_site = df_site.iloc[-self.ts:]

            self.train_x, self.train_y = \
                self.create_raster(df_site[:self.train_ts], self.train_ln, self.train_x,
                                   self.train_y)

            self.valid_x, self.valid_y = \
                self.create_raster(df_site[self.train_ts:self.train_ts+self.valid_ts], self.valid_ln,
                                   self.valid_x, self.valid_y)

        self.train_x = self.train_x[-params.max_train_len:, :, :]
        self.train_y = self.train_y[-params.max_train_len:, :, :]
        self.val_x = self.valid_x[:params.max_val_len, :, :]
        self.val_y = self.valid_y[:params.max_val_len:, :, :]
        self.test_x = self.valid_x[params.max_val_len:2*params.max_val_len, :, :]
        self.test_y = self.valid_y[params.max_val_len:2*params.max_val_len, :, :]

        pickle.dump(self.train_x, open("train_x.p", "wb"))
        pickle.dump(self.train_y, open("train_y.p", "wb"))
        pickle.dump(self.valid_x, open("valid_x.p", "wb"))
        pickle.dump(self.valid_y, open("valid_y.p", "wb"))
        pickle.dump(self.test_x, open("test_x.p", "wb"))
        pickle.dump(self.test_y, open("test_y.p", "wb"))
        pickle.dump(self.scalers, open("scalers.pkl", "wb"))

    def get_length(self, len):
        ln = len - (self.in_seq_len + self.out_seq_len)
        return ln

    def create_raster(self, data, ln, inputs, outputs):

        ts = len(data)

        stScaler = StandardScaler()
        dat = data.iloc[:, 2]
        dat = np.array(dat).reshape(-1, 1)
        stScaler.fit(dat)
        '''dat = stScaler.transform(dat)
        scaler.add_scaler(f, set_dat, stScaler)'''
        dat = torch.from_numpy(np.array(dat).flatten())
        in_data, out_data = self.get_window_data(dat, ln, ts)
        inputs[:, :, 0:1] = in_data
        outputs[:, :, 0] = out_data

        return inputs, outputs

    @staticmethod
    def moving_average(len, data, ts):

        data_mv = torch.zeros((ts, 1))
        for i in range(0, ts):
            if i < len:
                n = 1 if i == 0 else i
                data_mv[i, 0] = sum(data[0:i])/n
            else:
                data_mv[i, 0] = sum(data[i-len:i])/len
        return data_mv

    @staticmethod
    def get_derivative(k, data, ts):

        data_dv = torch.zeros((ts, 1))
        for i in range(0, ts):
            if i+k < ts:
                if i < k:
                    data_dv[i, 0] = data[i+k] - data[0]
                else:
                    data_dv[i, 0] = data[i+k] - data[i-k]
        return data_dv

    def create_wavelet(self, data, type):

        coeff = pywt.wavedec(data.detach().numpy(), type)
        arr, slices = pywt.coeffs_to_array(coeff)
        return torch.FloatTensor(arr[:data.shape[0]])

    def get_window_data(self, data, ln, ts):

        data_2d_in = torch.zeros((ts, self.n_moving_average+self.n_wavelets+1))
        data_3d_in = torch.zeros((ln, self.in_seq_len,
                                  self.n_moving_average+self.n_wavelets+1))
        data_out = torch.zeros((ln, self.out_seq_len))

        data_2d_in[:, 0] = data

        #print(data_2d_in)

        '''for i, mv in enumerate(self.moving_averages):

            data_2d_in[:, i] = self.moving_average(mv, data, ts).squeeze(1)
            #data_2d_in[:, i+self.n_moving_average] = self.get_derivative(mv, data, ts).squeeze(1)

        if self.n_wavelets > 0:
            for i, type in enumerate(self.wavelets):
                data_2d_in[:, self.n_moving_average+i:] = self.create_wavelet(data, type).unsqueeze(-1)'''

        j = 0
        for i in range(0, self.ts):
            if j < ln:
                data_3d_in[j, :, :] = data_2d_in[i:i+self.in_seq_len, :]
                data_out[j, :] = data[i+self.in_seq_len:i + self.in_seq_len + self.out_seq_len]
                j += 1
        return data_3d_in, data_out


class STData:
    def __init__(self, meta_path, site_path, params, site):
        self.meta_path = meta_path
        self.site_path = site_path
        self.I = 3
        self.J = 6
        self.n_features = 3
        self.sites_data = dict()
        site_dat = self.prep_data_per_site(site)
        self.sites_data[params.site] = site_dat
        self.raster = Data(self.sites_data, params,  self.n_features)

    class Site:
        def __init__(self, name, abr, lat, long):

            self.name = name
            self.abr = abr
            self.lat = lat
            self.long = long

    def move_near(self, grid, key, x, y):

        if y+1 < self.J and grid[x, y+1] is None:
            grid[x, y + 1] = key
        elif x+1 < self.I and grid[x+1, y] is None:
            grid[x + 1, y] = key
        elif x-1 >= 0 and grid[x-1, y] is None:
            grid[x - 1, y] = key
        elif y-1 >= 0 and grid[x, y-1] is None:
            grid[x, y - 1] = key

    def create_grid(self):

        wb_site = openpyxl.load_workbook(self.meta_path, data_only=True)
        sheet_site = wb_site.get_sheet_by_name("Site Info")
        max_row = sheet_site.max_row

        site_list = list()

        for row in range(2, max_row + 1):
            abr = sheet_site.cell(row=row, column=2).value
            name = sheet_site.cell(row=row, column=1).value
            lat = sheet_site.cell(row=row, column=10).value
            long = sheet_site.cell(row=row, column=11).value
            site_list.append(self.Site(name, abr, lat, long))

        min_lat, min_long, max_lat, max_long = 100, 100, -100, -100

        for site in site_list:

            lat, long = site.lat, site.long

            if lat < min_lat:
                min_lat = lat
            if lat > max_lat:
                max_lat = lat
            if long < min_long:
                min_long = long
            if long > max_long:
                max_long = long

        grid = None
        for i in range(self.I):
            ls = list()
            for j in range(self.J):
                ls.append(None)
            if grid is None:
                grid = ls
            else:
                grid = np.vstack((grid, ls))

        grid = np.array(grid)

        lat_diff = (max_lat - min_lat) / (self.I - 1)
        long_diff = (max_long - min_long) / (self.J - 1)

        for site in site_list:

            lat, long = site.lat, site.long
            x = int((lat - min_lat) / lat_diff)
            y = int((long - min_long) / long_diff)

            if grid[x, y] is None:
                grid[x, y] = site.abr
            else:
                self.move_near(grid, site.abr, x, y)

        grid_site = dict()

        for i in range(self.I):
            for j in range(self.J):
                if grid[i, j] is not None:
                    grid_site[grid[i, j]] = [i, j]

        return grid_site

    def prep_data_per_site(self, abr):

        df = pd.read_csv("{}/{}_WQual_Level4.csv".format(self.site_path, abr))
        df["Date"] = pd.to_datetime(df["Date"])
        df["Date"] = df["Date"].dt.normalize()
        start_date = "2012-12-14"
        end_date = "2019-01-02"
        mask = (df["Date"] >= start_date) & (df["Date"] <= end_date)
        df = df[mask]
        df = df[["Date", "TempC", "SpConductivity", "Q"]]
        '''plt.plot(np.arange(0, 1500), df.SpConductivity.iloc[-1500:], color='k')
        plt.tick_params(axis="x", bottom=False, top=False)
        plt.tick_params(axis="y", left=False, right=False)
        #plt.axis('off')
        plt.show()'''
        df = df.ffill().bfill()
        return df


def main():

    parser = argparse.ArgumentParser(description="preprocess argument parser")
    parser.add_argument("--in_seq_len", type=int, default=144)
    parser.add_argument("--out_seq_len", type=int, default=72)
    parser.add_argument("--site", type=str, default="BEF")
    parser.add_argument("--train_percent", type=float, default=0.8)
    parser.add_argument("--max_length", type=int, default=3200)
    parser.add_argument("--max_train_len", type=int, default=480)
    parser.add_argument("--max_val_len", type=int, default=60)
    parser.add_argument("--add_wave", type=str, default="False")
    params = parser.parse_args()
    stdata = STData("data/metadata.xlsx", "data", params, params.site)


if __name__ == '__main__':
    main()